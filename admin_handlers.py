from __future__ import annotations

import logging
from pathlib import Path

from aiogram import F, Router
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, FSInputFile, InlineKeyboardButton, InlineKeyboardMarkup, Message
from openpyxl import Workbook, load_workbook

import database as db
from config import ADMIN_PASSWORD, ADMIN_TG_IDS, DATA_DIR, MAX_BROADCAST_PER_RUN
from keyboards import (
    admin_broadcast_target_kb,
    admin_login_kb,
    admin_panel_kb,
    admin_report_actions_kb,
    admin_restore_kb,
    admin_task_actions_kb,
    admin_user_actions_kb,
)
from utils import ROLE_LABELS, event_card, h, report_card, response_card, task_card

router = Router()
logger = logging.getLogger(__name__)


class AdminStates(StatesGroup):
    waiting_password = State()
    waiting_reject_reason = State()
    waiting_ban_reason = State()
    waiting_broadcast_text = State()
    waiting_restore_file = State()
    waiting_search_query = State()


TABLE_EXPORTS = {
    'users': ['id', 'telegram_id', 'username', 'full_name', 'phone', 'role', 'city', 'specialization', 'about', 'rating', 'reviews_count', 'is_active', 'is_banned', 'banned_reason', 'created_at', 'updated_at'],
    'tasks': ['id', 'customer_id', 'city', 'category', 'title', 'description', 'budget', 'address', 'latitude', 'longitude', 'contact_text', 'status', 'selected_response_id', 'rejection_reason', 'views_count', 'responses_count', 'is_urgent', 'bumped_at', 'expires_at', 'created_at', 'updated_at'],
    'task_photos': ['id', 'task_id', 'file_id', 'created_at'],
    'task_responses': ['id', 'task_id', 'worker_id', 'message', 'offer_price', 'contact_text', 'status', 'created_at', 'updated_at'],
    'reviews': ['id', 'task_id', 'response_id', 'worker_id', 'customer_id', 'rating', 'text', 'created_at'],
    'favorites': ['id', 'worker_id', 'task_id', 'created_at'],
    'reports': ['id', 'reporter_id', 'task_id', 'target_user_id', 'reason', 'status', 'created_at', 'updated_at'],
    'task_events': ['id', 'task_id', 'actor_user_id', 'event_type', 'details', 'created_at'],
    'admin_actions': ['id', 'admin_tg_id', 'action', 'entity_type', 'entity_id', 'details', 'created_at'],
}


def is_admin(callback_or_message) -> bool:
    tg_id = callback_or_message.from_user.id
    return tg_id in ADMIN_TG_IDS or db.has_active_admin_session(tg_id)


async def admin_panel(message: Message | CallbackQuery):
    sender = message.message if isinstance(message, CallbackQuery) else message
    await sender.answer('🛠 <b>Админ-панель</b>', reply_markup=admin_panel_kb(), parse_mode='HTML')


@router.message(F.text == '🛠 Админ-панель')
@router.message(F.text == '/admin')
async def open_admin(message: Message, state: FSMContext):
    await state.clear()
    if is_admin(message):
        await admin_panel(message)
        return
    await message.answer('Введите пароль администратора:', reply_markup=admin_login_kb())
    await state.set_state(AdminStates.waiting_password)


@router.callback_query(F.data == 'admin_login')
async def admin_login(callback: CallbackQuery, state: FSMContext):
    if is_admin(callback):
        await admin_panel(callback)
        await callback.answer()
        return
    await callback.message.answer('Введите пароль администратора:')
    await state.set_state(AdminStates.waiting_password)
    await callback.answer()


@router.message(AdminStates.waiting_password)
async def admin_password(message: Message, state: FSMContext):
    if not ADMIN_PASSWORD:
        await message.answer('Пароль админа не задан в .env. Укажи ADMIN_PASSWORD или добавь свой Telegram ID в ADMIN_TG_IDS.')
        return
    if (message.text or '').strip() != ADMIN_PASSWORD:
        await message.answer('Неверный пароль.')
        return
    db.create_admin_session(message.from_user.id)
    db.log_admin_action(message.from_user.id, 'admin_login')
    await state.clear()
    await message.answer('Доступ открыт ✅')
    await admin_panel(message)


@router.callback_query(F.data == 'admin_panel')
async def admin_panel_callback(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    await admin_panel(callback)
    await callback.answer()


@router.callback_query(F.data == 'admin_logout')
async def admin_logout(callback: CallbackQuery):
    db.remove_admin_session(callback.from_user.id)
    db.log_admin_action(callback.from_user.id, 'admin_logout')
    await callback.message.answer('Выход из админки выполнен.')
    await callback.answer()


# ===== Tasks moderation =====

@router.callback_query(F.data == 'admin_pending')
async def admin_pending(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    tasks = db.get_tasks(status='pending', limit=30)
    if not tasks:
        await callback.message.answer('На модерации сейчас ничего нет.')
        await callback.answer()
        return
    buttons = [[InlineKeyboardButton(text=f'#{task["id"]} · {task["title"][:30]}', callback_data=f'admin_task_view:{task["id"]}')] for task in tasks]
    buttons.append([InlineKeyboardButton(text='🔙 Панель', callback_data='admin_panel')])
    await callback.message.answer('🕓 <b>Заявки на модерации</b>', reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons), parse_mode='HTML')
    await callback.answer()


@router.callback_query(F.data == 'admin_all_tasks')
async def admin_all_tasks(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    tasks = db.get_tasks(status='all', limit=30)
    if not tasks:
        await callback.message.answer('Заявок пока нет.')
        await callback.answer()
        return
    buttons = [[InlineKeyboardButton(text=f'#{task["id"]} · {task["title"][:22]} · {task["status"]}', callback_data=f'admin_task_view:{task["id"]}')] for task in tasks]
    buttons.append([InlineKeyboardButton(text='🔙 Панель', callback_data='admin_panel')])
    await callback.message.answer('📋 <b>Все заявки</b>', reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons), parse_mode='HTML')
    await callback.answer()


@router.callback_query(F.data.startswith('admin_task_view:'))
async def admin_task_view(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    task_id = int(callback.data.split(':', 1)[1])
    task = db.get_task(task_id)
    if not task:
        await callback.answer('Заявка не найдена', show_alert=True)
        return
    await callback.message.answer(task_card(task, full=True), reply_markup=admin_task_actions_kb(task_id, task['status']), parse_mode='HTML')
    await callback.answer()


@router.callback_query(F.data.startswith('admin_task_approve:'))
async def admin_task_approve(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    task_id = int(callback.data.split(':', 1)[1])
    task = db.get_task(task_id)
    if not task:
        await callback.answer('Не найдено', show_alert=True)
        return
    db.update_task_status(task_id, 'open')
    db.log_admin_action(callback.from_user.id, 'task_approved', 'task', task_id)
    db.log_task_event(task_id, 'approved', task['customer_id'], 'approved by admin')
    try:
        await callback.bot.send_message(task['customer_tg'], f'✅ Ваша заявка #{task_id} прошла модерацию и теперь видна исполнителям.')
    except Exception:
        logger.exception('Failed to notify customer about approval')
    await callback.message.answer('Заявка одобрена ✅')
    await callback.answer()


@router.callback_query(F.data.startswith('admin_task_reject:'))
async def admin_task_reject(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    task_id = int(callback.data.split(':', 1)[1])
    await state.update_data(reject_task_id=task_id)
    await state.set_state(AdminStates.waiting_reject_reason)
    await callback.message.answer('Напишите причину отклонения заявки.')
    await callback.answer()


@router.message(AdminStates.waiting_reject_reason)
async def admin_reject_reason(message: Message, state: FSMContext):
    if not is_admin(message):
        await message.answer('Нет доступа.')
        return
    data = await state.get_data()
    task_id = int(data['reject_task_id'])
    reason = (message.text or '').strip() or 'Без комментария'
    task = db.get_task(task_id)
    db.update_task_status(task_id, 'rejected', rejection_reason=reason)
    db.log_admin_action(message.from_user.id, 'task_rejected', 'task', task_id, reason)
    db.log_task_event(task_id, 'rejected', task['customer_id'] if task else None, reason)
    await state.clear()
    await message.answer('Заявка отклонена.')
    if task:
        try:
            await message.bot.send_message(task['customer_tg'], f'❌ Ваша заявка #{task_id} отклонена.\nПричина: {reason}')
        except Exception:
            logger.exception('Failed to notify customer about rejection')


@router.callback_query(F.data.startswith('admin_task_delete:'))
async def admin_task_delete(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    task_id = int(callback.data.split(':', 1)[1])
    db.delete_task(task_id)
    db.log_admin_action(callback.from_user.id, 'task_deleted', 'task', task_id)
    await callback.message.answer(f'Заявка #{task_id} удалена.')
    await callback.answer()


@router.callback_query(F.data.startswith('admin_task_bump:'))
async def admin_task_bump(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    task_id = int(callback.data.split(':', 1)[1])
    ok, msg = db.bump_task(task_id, urgent=True)
    db.log_admin_action(callback.from_user.id, 'task_bumped', 'task', task_id, msg)
    await callback.message.answer(('✅ ' if ok else '⚠️ ') + msg)
    await callback.answer()


@router.callback_query(F.data.startswith('admin_task_responses:'))
async def admin_task_responses(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    task_id = int(callback.data.split(':', 1)[1])
    responses = db.get_task_responses(task_id)
    if not responses:
        await callback.message.answer('По заявке пока нет откликов.')
        await callback.answer()
        return
    for response in responses[:20]:
        await callback.message.answer(response_card(response), parse_mode='HTML')
    await callback.answer()


@router.callback_query(F.data.startswith('admin_task_events:'))
async def admin_task_events(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    task_id = int(callback.data.split(':', 1)[1])
    events = db.get_task_events(task_id, limit=20)
    if not events:
        await callback.message.answer('История пуста.')
    else:
        for event in events:
            await callback.message.answer(event_card(event), parse_mode='HTML')
    await callback.answer()


# ===== Users =====

@router.callback_query(F.data == 'admin_users')
async def admin_users(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    await _send_role_list(callback.message, 'customer', '👥 Заказчики')
    await callback.answer()


@router.callback_query(F.data == 'admin_workers')
async def admin_workers(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    await _send_role_list(callback.message, 'worker', '🧑‍🔧 Исполнители')
    await callback.answer()


@router.callback_query(F.data.startswith('admin_list_role:'))
async def admin_list_role(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    role = callback.data.split(':', 1)[1]
    title = '👥 Пользователи' if role == 'customer' else '🧑‍🔧 Исполнители'
    await _send_role_list(callback.message, role, title)
    await callback.answer()


async def _send_role_list(message: Message, role: str, title: str):
    users = db.list_users(role=role, limit=50)
    if not users:
        await message.answer('Список пуст.')
        return
    buttons = [[InlineKeyboardButton(text=f'#{user["id"]} · {user["full_name"][:30]}', callback_data=f'admin_user_view:{user["id"]}')] for user in users]
    buttons.append([InlineKeyboardButton(text='🔙 Панель', callback_data='admin_panel')])
    await message.answer(title, reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))


@router.callback_query(F.data.startswith('admin_user_view:'))
async def admin_user_view(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    user_id = int(callback.data.split(':', 1)[1])
    user = db.get_user_by_id(user_id)
    if not user:
        await callback.answer('Пользователь не найден', show_alert=True)
        return
    text = (
        f'👤 <b>Пользователь #{user["id"]}</b>\n\n'
        f'Имя: {h(user["full_name"])}\n'
        f'Роль: {ROLE_LABELS.get(user["role"], user["role"])}\n'
        f'Город: {h(user.get("city") or "—")}\n'
        f'Телефон: {h(user.get("phone") or "—")}\n'
        f'Username: @{h(user.get("username") or "—")}\n'
        f'Рейтинг: {user.get("rating") or 0} ({user.get("reviews_count") or 0} отзывов)\n'
        f'Бан: {"да" if user.get("is_banned") else "нет"}\n'
    )
    if user.get('specialization'):
        text += f'Специализация: {h(user.get("specialization"))}\n'
    if user.get('about'):
        text += f'О себе: {h(user.get("about"))}\n'
    await callback.message.answer(text, reply_markup=admin_user_actions_kb(user['id'], bool(user.get('is_banned')), user['role']), parse_mode='HTML')
    await callback.answer()


@router.callback_query(F.data.startswith('admin_ban:'))
async def admin_ban(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    user_id = int(callback.data.split(':', 1)[1])
    await state.update_data(ban_user_id=user_id)
    await state.set_state(AdminStates.waiting_ban_reason)
    await callback.message.answer('Напиши причину бана.')
    await callback.answer()


@router.message(AdminStates.waiting_ban_reason)
async def admin_ban_reason(message: Message, state: FSMContext):
    if not is_admin(message):
        await message.answer('Нет доступа.')
        return
    data = await state.get_data()
    user_id = int(data['ban_user_id'])
    reason = (message.text or '').strip() or 'Нарушение правил'
    user = db.get_user_by_id(user_id)
    db.ban_user(user_id, reason)
    db.log_admin_action(message.from_user.id, 'user_banned', 'user', user_id, reason)
    await state.clear()
    await message.answer('Пользователь заблокирован.')
    if user:
        try:
            await message.bot.send_message(user['telegram_id'], f'🚫 Ваш аккаунт заблокирован. Причина: {reason}')
        except Exception:
            logger.exception('Failed to notify banned user')


@router.callback_query(F.data.startswith('admin_unban:'))
async def admin_unban(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    user_id = int(callback.data.split(':', 1)[1])
    user = db.get_user_by_id(user_id)
    db.unban_user(user_id)
    db.log_admin_action(callback.from_user.id, 'user_unbanned', 'user', user_id)
    await callback.message.answer('Пользователь разбанен.')
    if user:
        try:
            await callback.bot.send_message(user['telegram_id'], '✅ Ваш аккаунт разблокирован.')
        except Exception:
            logger.exception('Failed to notify unbanned user')
    await callback.answer()


@router.callback_query(F.data.startswith('admin_user_reports:'))
async def admin_user_reports(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    user_id = int(callback.data.split(':', 1)[1])
    reports = [r for r in db.get_reports(status='all', limit=100) if r.get('target_user_id') == user_id]
    if not reports:
        await callback.message.answer('Жалоб на этого пользователя нет.')
    else:
        for report in reports[:20]:
            await callback.message.answer(report_card(report), parse_mode='HTML')
    await callback.answer()


# ===== Reports =====

@router.callback_query(F.data == 'admin_reports')
async def admin_reports(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    reports = db.get_reports(status='open', limit=30)
    if not reports:
        await callback.message.answer('Открытых жалоб нет.')
        await callback.answer()
        return
    for report in reports:
        await callback.message.answer(report_card(report), reply_markup=admin_report_actions_kb(report['id'], report['status']), parse_mode='HTML')
    await callback.answer()


@router.callback_query(F.data.startswith('admin_report_close:'))
async def admin_report_close(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    report_id = int(callback.data.split(':', 1)[1])
    db.close_report(report_id)
    db.log_admin_action(callback.from_user.id, 'report_closed', 'report', report_id)
    await callback.message.answer(f'Жалоба #{report_id} закрыта.')
    await callback.answer()


# ===== Search =====

@router.callback_query(F.data == 'admin_search')
async def admin_search(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    await state.set_state(AdminStates.waiting_search_query)
    await callback.message.answer('Введи ID заявки, ID/Telegram ID пользователя, имя, username или часть текста заявки.')
    await callback.answer()


@router.message(AdminStates.waiting_search_query)
async def admin_search_query(message: Message, state: FSMContext):
    if not is_admin(message):
        await message.answer('Нет доступа.')
        return
    query = (message.text or '').strip()
    if not query:
        await message.answer('Пустой запрос.')
        return
    users = db.search_users(query, limit=10)
    tasks = db.search_tasks(query, status='all', limit=10)
    await state.clear()
    if not users and not tasks:
        await message.answer('Ничего не найдено.')
        return
    if tasks:
        buttons = [[InlineKeyboardButton(text=f'Заявка #{t["id"]} · {t["title"][:22]}', callback_data=f'admin_task_view:{t["id"]}')] for t in tasks]
        await message.answer('🔎 <b>Найденные заявки</b>', reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons), parse_mode='HTML')
    if users:
        buttons = [[InlineKeyboardButton(text=f'Пользователь #{u["id"]} · {u["full_name"][:24]}', callback_data=f'admin_user_view:{u["id"]}')] for u in users]
        await message.answer('🔎 <b>Найденные пользователи</b>', reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons), parse_mode='HTML')


# ===== Stats =====

@router.callback_query(F.data == 'admin_stats')
async def admin_stats(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    stats = db.get_stats()
    text = (
        '📊 <b>Статистика</b>\n\n'
        f'Пользователи: {stats["users_total"]}\n'
        f'Заказчики: {stats["customers_total"]}\n'
        f'Исполнители: {stats["workers_total"]}\n'
        f'В бане: {stats["banned_total"]}\n\n'
        f'Всего заявок: {stats["tasks_total"]}\n'
        f'На модерации: {stats["tasks_pending"]}\n'
        f'Открыты: {stats["tasks_open"]}\n'
        f'В работе: {stats["tasks_in_progress"]}\n'
        f'Закрыты: {stats["tasks_closed"]}\n'
        f'В архиве: {stats["tasks_archived"]}\n\n'
        f'Всего откликов: {stats["responses_total"]}\n'
        f'Избранное: {stats["favorites_total"]}\n'
        f'Открытые жалобы: {stats["reports_open"]}\n'
        f'Средний бюджет: {int(stats["avg_budget"] or 0)} ₽'
    )
    await callback.message.answer(text, parse_mode='HTML')
    await callback.answer()


# ===== Broadcast =====

@router.callback_query(F.data == 'admin_broadcast')
async def admin_broadcast(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    await callback.message.answer('Выбери, кому отправлять:', reply_markup=admin_broadcast_target_kb())
    await callback.answer()


@router.callback_query(F.data.startswith('broadcast_target:'))
async def broadcast_target(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    target = callback.data.split(':', 1)[1]
    await state.update_data(broadcast_target=target)
    await state.set_state(AdminStates.waiting_broadcast_text)
    await callback.message.answer('Отправь текст рассылки.')
    await callback.answer()


@router.message(AdminStates.waiting_broadcast_text)
async def broadcast_send(message: Message, state: FSMContext):
    if not is_admin(message):
        await message.answer('Нет доступа.')
        return
    text = (message.text or '').strip()
    if not text:
        await message.answer('Текст пустой.')
        return
    data = await state.get_data()
    target = data.get('broadcast_target', 'all')
    users = db.list_users(role=target if target in {'customer', 'worker'} else None, limit=MAX_BROADCAST_PER_RUN)
    ok = 0
    fail = 0
    for user in users:
        try:
            await message.bot.send_message(user['telegram_id'], text)
            ok += 1
        except Exception:
            fail += 1
    db.log_admin_action(message.from_user.id, 'broadcast', details=f'target={target}; ok={ok}; fail={fail}')
    await state.clear()
    await message.answer(f'Рассылка завершена. Успешно: {ok}, ошибок: {fail}.')


# ===== Export DB / Excel =====

@router.callback_query(F.data == 'admin_export_db')
async def admin_export_db(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    target = DATA_DIR / 'backup_job_bot_v4.db'
    db.backup_database_copy(str(target))
    await callback.message.answer_document(FSInputFile(target), caption='💾 Резервная копия базы')
    db.log_admin_action(callback.from_user.id, 'export_db')
    await callback.answer()


@router.callback_query(F.data == 'admin_export_excel')
async def admin_export_excel(callback: CallbackQuery):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    file_path = DATA_DIR / 'job_bot_v4_export.xlsx'
    wb = Workbook()
    first = True
    for sheet_name, headers in TABLE_EXPORTS.items():
        ws = wb.active if first else wb.create_sheet(title=sheet_name)
        ws.title = sheet_name
        first = False
        ws.append(headers)
        for row in db.get_all_table_rows(sheet_name):
            ws.append([row.get(col) for col in headers])
    wb.save(file_path)
    await callback.message.answer_document(FSInputFile(file_path), caption='📊 Excel-выгрузка')
    db.log_admin_action(callback.from_user.id, 'export_excel')
    await callback.answer()


# ===== Restore =====

@router.callback_query(F.data == 'admin_restore')
async def admin_restore(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    await state.clear()
    await callback.message.answer('Выбери тип восстановления:', reply_markup=admin_restore_kb())
    await callback.answer()


@router.callback_query(F.data.in_({'admin_restore_db', 'admin_restore_excel_file'}))
async def admin_restore_mode(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback):
        await callback.answer('Нет доступа', show_alert=True)
        return
    mode = 'db' if callback.data == 'admin_restore_db' else 'excel'
    await state.update_data(restore_mode=mode)
    await state.set_state(AdminStates.waiting_restore_file)
    await callback.message.answer('Теперь отправь файл для восстановления.')
    await callback.answer()


@router.message(AdminStates.waiting_restore_file, F.document)
async def admin_restore_file(message: Message, state: FSMContext):
    if not is_admin(message):
        await message.answer('Нет доступа.')
        return
    data = await state.get_data()
    mode = data.get('restore_mode')
    suffix = '.db' if mode == 'db' else '.xlsx'
    if not (message.document.file_name or '').lower().endswith(suffix):
        await message.answer(f'Нужен файл формата {suffix}')
        return
    temp_path = DATA_DIR / f'restore_upload{suffix}'
    await message.bot.download(message.document, destination=temp_path)
    try:
        if mode == 'db':
            db.restore_database_from_copy(str(temp_path))
        else:
            payload = _read_excel_payload(temp_path)
            db.replace_from_excel_rows(**payload)
        db.log_admin_action(message.from_user.id, 'restore_db', details=mode)
        await message.answer('Восстановление завершено ✅')
    except Exception as e:
        logger.exception('Restore failed')
        await message.answer(f'Ошибка восстановления: {e}')
    finally:
        await state.clear()


@router.message(AdminStates.waiting_restore_file)
async def admin_restore_invalid(message: Message):
    await message.answer('Жду файл .db или .xlsx в зависимости от выбранного режима.')


# ===== Helpers =====

def _read_excel_payload(path: Path) -> dict[str, list[dict]]:
    wb = load_workbook(path)
    payload: dict[str, list[dict]] = {}
    for sheet_name, headers in TABLE_EXPORTS.items():
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            payload[sheet_name] = []
            continue
        header = [str(x) for x in rows[0]]
        data_rows = []
        for row in rows[1:]:
            item = {}
            empty = True
            for idx, key in enumerate(header):
                value = row[idx] if idx < len(row) else None
                if value is not None:
                    empty = False
                item[key] = value
            if not empty:
                data_rows.append(item)
        payload[sheet_name] = data_rows
    return payload
